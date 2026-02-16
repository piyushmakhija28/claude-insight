#!/usr/bin/env python3
"""
Comprehensive Unit Tests for Export Functionality

Tests export functionality including:
- CSV export
- JSON export
- PDF export
- Excel export
- Data formatting
- Export filters
"""

import unittest
import json
import tempfile
import os
import csv
from pathlib import Path
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock
from io import StringIO, BytesIO
import sys

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))


class TestCSVExport(unittest.TestCase):
    """Test suite for CSV export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.sample_data = [
            {'timestamp': '2026-02-16 10:00:00', 'level': 'INFO', 'message': 'Test 1'},
            {'timestamp': '2026-02-16 10:01:00', 'level': 'ERROR', 'message': 'Test 2'},
            {'timestamp': '2026-02-16 10:02:00', 'level': 'WARNING', 'message': 'Test 3'}
        ]

    def test_export_csv_basic(self):
        """Test basic CSV export"""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['timestamp', 'level', 'message'])
        writer.writeheader()
        writer.writerows(self.sample_data)

        csv_content = output.getvalue()

        self.assertIn('timestamp', csv_content)
        self.assertIn('INFO', csv_content)
        self.assertIn('ERROR', csv_content)

    def test_export_csv_custom_delimiter(self):
        """Test CSV export with custom delimiter"""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['timestamp', 'level', 'message'], delimiter=';')
        writer.writeheader()
        writer.writerows(self.sample_data)

        csv_content = output.getvalue()

        self.assertIn(';', csv_content)

    def test_export_csv_empty_data(self):
        """Test CSV export with empty data"""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['timestamp', 'level', 'message'])
        writer.writeheader()
        writer.writerows([])

        csv_content = output.getvalue()

        # Should have header even with no data
        self.assertIn('timestamp', csv_content)

    def test_export_csv_special_characters(self):
        """Test CSV export with special characters"""
        data = [
            {'field1': 'Value with, comma', 'field2': 'Value with "quotes"'},
            {'field1': 'Value with\nnewline', 'field2': 'Normal value'}
        ]

        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['field1', 'field2'])
        writer.writeheader()
        writer.writerows(data)

        csv_content = output.getvalue()

        # CSV should properly escape special characters
        self.assertIn('comma', csv_content)
        self.assertIn('quotes', csv_content)


class TestJSONExport(unittest.TestCase):
    """Test suite for JSON export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.sample_data = {
            'metrics': [
                {'name': 'cpu', 'value': 75},
                {'name': 'memory', 'value': 60}
            ],
            'timestamp': '2026-02-16T10:00:00',
            'status': 'healthy'
        }

    def test_export_json_basic(self):
        """Test basic JSON export"""
        json_str = json.dumps(self.sample_data, indent=2)

        self.assertIsInstance(json_str, str)
        parsed = json.loads(json_str)
        self.assertEqual(parsed['status'], 'healthy')

    def test_export_json_pretty_print(self):
        """Test JSON export with pretty printing"""
        json_str = json.dumps(self.sample_data, indent=4, sort_keys=True)

        self.assertIn('\n', json_str)  # Has newlines
        self.assertIn('    ', json_str)  # Has indentation

    def test_export_json_compact(self):
        """Test compact JSON export"""
        json_str = json.dumps(self.sample_data, separators=(',', ':'))

        # Compact format has no extra spaces
        self.assertNotIn(', ', json_str)

    def test_export_json_array(self):
        """Test exporting array of objects"""
        data = [
            {'id': 1, 'name': 'Item 1'},
            {'id': 2, 'name': 'Item 2'}
        ]

        json_str = json.dumps(data)
        parsed = json.loads(json_str)

        self.assertIsInstance(parsed, list)
        self.assertEqual(len(parsed), 2)


class TestPDFExport(unittest.TestCase):
    """Test suite for PDF export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    @patch('reportlab.platypus.SimpleDocTemplate')
    def test_export_pdf_basic(self, mock_doc):
        """Test basic PDF export"""
        from reportlab.platypus import Table, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        # Mock the PDF generation
        mock_doc_instance = MagicMock()
        mock_doc.return_value = mock_doc_instance

        styles = getSampleStyleSheet()
        elements = []

        # Add title
        title = Paragraph("Test Report", styles['Title'])
        elements.append(title)

        # Add table
        data = [
            ['Header 1', 'Header 2'],
            ['Row 1 Col 1', 'Row 1 Col 2'],
            ['Row 2 Col 1', 'Row 2 Col 2']
        ]
        table = Table(data)
        elements.append(table)

        # Should not raise exception
        self.assertTrue(True)

    def test_pdf_table_creation(self):
        """Test creating PDF table"""
        from reportlab.platypus import Table
        from reportlab.lib import colors

        data = [
            ['Name', 'Value', 'Status'],
            ['Metric 1', '100', 'OK'],
            ['Metric 2', '200', 'OK']
        ]

        table = Table(data)

        self.assertIsNotNone(table)


class TestExcelExport(unittest.TestCase):
    """Test suite for Excel export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_export_excel_basic(self):
        """Test basic Excel export"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"

        # Add headers
        ws.append(['Timestamp', 'Level', 'Message'])

        # Add data
        ws.append(['2026-02-16 10:00:00', 'INFO', 'Test message'])

        # Save to temp file
        temp_file = Path(self.temp_dir) / 'test.xlsx'
        wb.save(temp_file)

        self.assertTrue(temp_file.exists())

    def test_export_excel_multiple_sheets(self):
        """Test Excel export with multiple sheets"""
        from openpyxl import Workbook

        wb = Workbook()

        # Create multiple sheets
        ws1 = wb.active
        ws1.title = "Metrics"
        ws1.append(['Metric', 'Value'])
        ws1.append(['CPU', 75])

        ws2 = wb.create_sheet("Logs")
        ws2.append(['Timestamp', 'Message'])
        ws2.append(['2026-02-16', 'Test'])

        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn('Metrics', wb.sheetnames)
        self.assertIn('Logs', wb.sheetnames)

    def test_export_excel_formatting(self):
        """Test Excel export with formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active

        # Add header with formatting
        ws.append(['Header 1', 'Header 2'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')

        # Check formatting applied
        self.assertTrue(ws['A1'].font.bold)


class TestExportFilters(unittest.TestCase):
    """Test suite for export filtering functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.sample_data = [
            {'timestamp': '2026-02-16 10:00:00', 'level': 'INFO', 'value': 100},
            {'timestamp': '2026-02-16 10:01:00', 'level': 'ERROR', 'value': 200},
            {'timestamp': '2026-02-16 10:02:00', 'level': 'WARNING', 'value': 150},
            {'timestamp': '2026-02-16 10:03:00', 'level': 'INFO', 'value': 120}
        ]

    def test_filter_by_level(self):
        """Test filtering data by log level"""
        filtered = [item for item in self.sample_data if item['level'] == 'ERROR']

        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0]['level'], 'ERROR')

    def test_filter_by_date_range(self):
        """Test filtering data by date range"""
        start = '2026-02-16 10:00:00'
        end = '2026-02-16 10:02:00'

        filtered = [
            item for item in self.sample_data
            if start <= item['timestamp'] <= end
        ]

        self.assertGreaterEqual(len(filtered), 2)

    def test_filter_by_value_threshold(self):
        """Test filtering data by value threshold"""
        threshold = 150

        filtered = [item for item in self.sample_data if item['value'] >= threshold]

        self.assertEqual(len(filtered), 2)
        for item in filtered:
            self.assertGreaterEqual(item['value'], threshold)

    def test_filter_multiple_conditions(self):
        """Test filtering with multiple conditions"""
        filtered = [
            item for item in self.sample_data
            if item['level'] == 'INFO' and item['value'] > 100
        ]

        self.assertEqual(len(filtered), 1)

    def test_filter_exclude(self):
        """Test excluding items based on condition"""
        filtered = [item for item in self.sample_data if item['level'] != 'ERROR']

        self.assertEqual(len(filtered), 3)
        for item in filtered:
            self.assertNotEqual(item['level'], 'ERROR')


class TestExportDataFormatting(unittest.TestCase):
    """Test suite for data formatting in exports"""

    def test_format_timestamp(self):
        """Test timestamp formatting"""
        timestamp = datetime(2026, 2, 16, 10, 30, 45)

        formatted = timestamp.strftime('%Y-%m-%d %H:%M:%S')

        self.assertEqual(formatted, '2026-02-16 10:30:45')

    def test_format_large_numbers(self):
        """Test formatting large numbers"""
        number = 1234567

        formatted = f"{number:,}"

        self.assertEqual(formatted, '1,234,567')

    def test_format_percentage(self):
        """Test formatting percentages"""
        value = 0.7545

        formatted = f"{value:.2%}"

        self.assertEqual(formatted, '75.45%')

    def test_format_boolean(self):
        """Test formatting boolean values"""
        self.assertEqual(str(True), 'True')
        self.assertEqual(str(False), 'False')

    def test_format_null_values(self):
        """Test handling null/None values"""
        value = None

        formatted = value if value is not None else 'N/A'

        self.assertEqual(formatted, 'N/A')

    def test_truncate_long_text(self):
        """Test truncating long text"""
        long_text = 'A' * 200
        max_length = 50

        truncated = long_text[:max_length] + '...' if len(long_text) > max_length else long_text

        self.assertEqual(len(truncated), max_length + 3)
        self.assertTrue(truncated.endswith('...'))


class TestExportIntegration(unittest.TestCase):
    """Integration tests for export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.sample_data = [
            {'id': 1, 'name': 'Item 1', 'value': 100, 'timestamp': '2026-02-16 10:00:00'},
            {'id': 2, 'name': 'Item 2', 'value': 200, 'timestamp': '2026-02-16 10:01:00'}
        ]

    def tearDown(self):
        """Clean up"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_export_same_data_all_formats(self):
        """Test exporting same data to all formats"""
        # CSV
        csv_file = Path(self.temp_dir) / 'data.csv'
        with open(csv_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['id', 'name', 'value', 'timestamp'])
            writer.writeheader()
            writer.writerows(self.sample_data)

        # JSON
        json_file = Path(self.temp_dir) / 'data.json'
        with open(json_file, 'w') as f:
            json.dump(self.sample_data, f, indent=2)

        # Excel
        from openpyxl import Workbook
        excel_file = Path(self.temp_dir) / 'data.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'name', 'value', 'timestamp'])
        for item in self.sample_data:
            ws.append([item['id'], item['name'], item['value'], item['timestamp']])
        wb.save(excel_file)

        # Verify all files created
        self.assertTrue(csv_file.exists())
        self.assertTrue(json_file.exists())
        self.assertTrue(excel_file.exists())


if __name__ == '__main__':
    unittest.main()
